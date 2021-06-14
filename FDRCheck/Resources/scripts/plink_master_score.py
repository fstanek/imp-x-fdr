import xlrd, xlsxwriter
import sys, os, csv
from openpyxl import Workbook

# open the sheetspreadfile and the respective sheet 
workbook = xlrd.open_workbook('generate_xl_lys_test.xlsx', on_demand = True)
worksheet = workbook.sheet_by_name('Sheet1')


#list_crosslinks_connect_acid = ["D","E"]
#list_crosslinks_connect_lys = ["K","Y","T","S"]

input_filename = sys.argv[1]
output_filename = sys.argv[2]
#number_groups = int(input("Please type in the number of peptide groups: "))
#peptides_per_group = int(input("Please type in the number of peptides per group: "))
number_groups = sys.argv[3]
peptides_per_group = sys.argv[4]

# in list_of_groups there will be all peptides groups with all their members stored 
list_of_groups = []
# will be added to the list_of_groups and contains temporary the members of one group at the time
list_of_peptides_per_group = []


# all peptide sequences present in the excel document are read one by oneand the added to the list of groups
# list of peptides per group will be permanently emptied to make space for the other groups 
# the number of list of peptides per group is static, not dynamic and should be kept constant
for i in range(1,number_groups+1):
    for j in range(1,peptides_per_group+1):
        value = worksheet.cell(((i-1)*(peptides_per_group+1))+j, 0).value
        list_of_peptides_per_group.append(value)
    a = list_of_peptides_per_group
    list_of_groups.append(a)
    list_of_peptides_per_group = []


counter=0
wb = Workbook()
ws = wb.active
with open(input_filename, 'r') as f:
    for row in csv.reader(f, delimiter='§'):
        ws.append(row)
        counter=counter+1


wb.save('plink_bin.xlsx')

wb = xlrd.open_workbook('plink_bin.xlsx', on_demand = True)
ws = wb.sheet_by_name('Sheet')

csm_crosslinks = []
list_of_scores_plink = []

for i in range(1,counter):
    pivot = ws.cell(i,0).value.split(",")
    score_plink = float(pivot[11])
    unprocessed_peptide_sequences = pivot[5]
    unprocessed_protein_positions = pivot[14]


    stop_position_1 = unprocessed_peptide_sequences.find('(')
    peptide1 = unprocessed_peptide_sequences[0:stop_position_1]

    start_position_2= unprocessed_peptide_sequences.find('-')
    unprocessed_peptide_sequences = unprocessed_peptide_sequences[start_position_2+1:]
    stop_position_2= unprocessed_peptide_sequences.find('(')
    peptide2= unprocessed_peptide_sequences[0:stop_position_2]


    start_protein_1 = unprocessed_protein_positions.find('sg|')+3
    unprocessed_protein_positions=unprocessed_protein_positions[start_protein_1:]
    stop_protein_1 = unprocessed_protein_positions.find('|')
    protein_1 = unprocessed_protein_positions[0:stop_protein_1]
    
    start_position_protein_1 = unprocessed_protein_positions.find('(')+1
    stop_position_protein_1 = unprocessed_protein_positions.find(')')
    try:
        protein_position_1 = int(str(unprocessed_protein_positions[start_position_protein_1:stop_position_protein_1]))
    except:
        if(start_position_protein_1==stop_position_protein_1):
            protein_position_1 = int(unprocessed_protein_positions[start_position_protein_1])
       
    unprocessed_protein_positions = unprocessed_protein_positions[stop_position_protein_1:]
    start_protein_2 = unprocessed_protein_positions.find('|')+1
    
    unprocessed_protein_positions = unprocessed_protein_positions[start_protein_2:]
    stop_protein_2 = unprocessed_protein_positions.find('|')
    protein_2 = unprocessed_protein_positions[0:stop_protein_2]
    try:
        protein_position_2 = int(str(unprocessed_protein_positions[unprocessed_protein_positions.find('(')+1:unprocessed_protein_positions.find(')')]))
    except:
        if(unprocessed_protein_positions.find('(')+1==unprocessed_protein_positions.find(')')-1):
            protein_position_2=int(unprocessed_protein_positions[unprocessed_protein_positions.find('(')+1])
    csm_crosslinks.append([peptide1,peptide2,score_plink,protein_1,protein_2,protein_position_1,protein_position_2])
    list_of_scores_plink.append(score_plink)


def order_alphabetically(example_list):

    example_list2= []

    for i in range(len(example_list)):
        example_list2 = example_list[i][0:2]

        
        if (example_list[i][0:2] != sorted(example_list2[0:2])):
            
            exchange = example_list[i][0]
            example_list[i][0] = example_list[i][1]
            example_list[i][1] = exchange

            exchange = example_list[i][3]
            example_list[i][3] = example_list[i][4]
            example_list[i][4] = exchange

            exchange = example_list[i][5]
            example_list[i][5] = example_list[i][6]
            example_list[i][6] = exchange


            

    return  example_list;


csm_crosslinks = order_alphabetically(csm_crosslinks)

list_of_scores_plink = list(set(list_of_scores_plink))
list_of_scores_plink.sort()

import itertools
csm_crosslinks.sort()
csm_crosslinks = list(csm_crosslinks for csm_crosslinks,_ in itertools.groupby(csm_crosslinks))


unique_crosslinks = []
unique_crosslinks.append(csm_crosslinks[0])
counter_xl = 0

print(list_of_scores_plink)
for i in range(1,len(csm_crosslinks)):
    if (unique_crosslinks[counter_xl][0]==csm_crosslinks[i][0] and
        unique_crosslinks[counter_xl][1]==csm_crosslinks[i][1] and
        unique_crosslinks[counter_xl][3]==csm_crosslinks[i][3] and
        unique_crosslinks[counter_xl][4]==csm_crosslinks[i][4] and
        unique_crosslinks[counter_xl][5]==csm_crosslinks[i][5] and
        unique_crosslinks[counter_xl][6]==csm_crosslinks[i][6] ):
        do_nothing = 0
    else:
        unique_crosslinks.append(csm_crosslinks[i])
        counter_xl=counter_xl+1
import math
for i in range(len(unique_crosslinks)):
    unique_crosslinks[i][2] = -(math.log(unique_crosslinks[i][2],10))

for i in range(len(list_of_scores_plink)):
    list_of_scores_plink[i] = -(math.log(list_of_scores_plink[i],10))
list_of_scores_plink.sort()

print(len(unique_crosslinks))


#fdr_cutoff_value = input("please type in the fdr cutoff value (ex:0.05): ")
fdr_cutoff_value = 0.05
temp1 = []
temp2 = []

def fdr_diagamm(list_crosslinks):
    temp1 = []
    temp2 = []

    import xlsxwriter

    workbook = xlsxwriter.Workbook(output_filename)
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0

    worksheet.write(0,0, "Results")
    worksheet.write(2,0, "total number of CSMs:")
    worksheet.write(3,0, "total number of unique XLs:")
    worksheet.write(4,0, "all True crosslinks: ")
    worksheet.write(5,0, "homeotypic crosslinks:")
    worksheet.write(6,0, "non-homeotypic crosslinks:")
    worksheet.write(7,0, "false crosslinks")

    worksheet.write(9,0, "all crosslinks")
    worksheet.write(9,1, "true crosslinks")
    worksheet.write(9,2, "false crosslinks")


    worksheet.set_column(0,0,30)

    

    



    to_be_ploted_x = []
    to_be_ploted_y = []
    correct_no_homo_XL = []
    homo_XL = []
    false_XL = []
    
    import matplotlib.pyplot as plt


    for i in range(len(list_of_scores_plink)):
        all_crosslinks = 0
        correct_crosslinks = 0
        homeotypic = 0
        for j in range(len(unique_crosslinks)):
            if unique_crosslinks[j][2]>=list_of_scores_plink[i]:
                all_crosslinks = all_crosslinks +1
                temp1.append(list([unique_crosslinks[j][0],unique_crosslinks[j][1],
                                   unique_crosslinks[j][3],unique_crosslinks[j][4],
                                   str(unique_crosslinks[j][5]),str(unique_crosslinks[j][6]),str("score_" + str(unique_crosslinks[j][2]))]))
                found = False
                k = 0
                while(k<number_groups and found == False):
                    l = 0
                    while(l<peptides_per_group and found == False):
                        if unique_crosslinks[j][0] == list_of_groups[k][l] or unique_crosslinks[j][0] in list_of_groups[k][l] :
                            m = 0
                            while(m<peptides_per_group and found== False):
                                if (unique_crosslinks[j][1]== list_of_groups[k][m] or unique_crosslinks[j][1] in list_of_groups[k][m]):
                                    found = True
                                    correct_crosslinks = correct_crosslinks +1
                                    temp2.append(list([unique_crosslinks[j][0],unique_crosslinks[j][1],
                                                       unique_crosslinks[j][3],unique_crosslinks[j][4],
                                                       str(unique_crosslinks[j][5]),str(unique_crosslinks[j][6]),str("score_" + str(unique_crosslinks[j][2]))]))
                                    if unique_crosslinks[j][0] == unique_crosslinks[j][1]:
                                        homeotypic = homeotypic+1
                                else:
                                    m = m+1
                            
                        l = l+1
                    k = k+1
        if i == 0:
            temp11 = []
            temp22 = []

            for m in range(len(temp1)):
                temp1[m].sort()
                temp1[m]=str(temp1[m])
                temp11.append(str(temp1[m]))
                worksheet.write(11+m,0,str(temp1[m]))


            for m in range(len(temp2)):
                temp2[m].sort()
                temp2[m]=str(temp2[m])
                temp22.append(str(temp2[m]))
                worksheet.write(11+m,1,str(temp2[m]))

            temp3=list(set(temp11)-set(temp22)) 
            for m in range(len(temp3)):
                worksheet.write(11+m,2,str(temp3[m]))
            


            workbook.close()
            list_of_the_correct_crosslinks = set(temp2)
            list_of_the_false_crosslinks = set(set(temp11)-set(temp22))
            list_of_all_crosslinks = set(temp1)



        #print(all_crosslinks-correct_crosslinks, all_crosslinks)
        to_be_ploted_x.append(list_of_scores_plink[i])
        to_be_ploted_y.append((all_crosslinks-correct_crosslinks)/all_crosslinks)
        correct_no_homo_XL.append(correct_crosslinks-homeotypic)
        homo_XL.append(homeotypic)
        false_XL.append(all_crosslinks-correct_crosslinks)
        
    
    plt.plot(to_be_ploted_x,to_be_ploted_y)

    plt.title('Real FDR dependent of the score')
    plt.xlabel('-log(Score)')
    plt.ylabel('FDR')
    alarm_005 = False
    alarm_001 = False

    bar_graph = ["real FDR"+" " +str(float("{0:.1f}".format((false_XL[0]/(correct_no_homo_XL[0]+homo_XL[0]+false_XL[0]))*100)))+"%"]
    
    gold = [correct_no_homo_XL[0]]
    silver = [homo_XL[0]]
    bronze = [false_XL[0]]


    plt.scatter(to_be_ploted_x[0],to_be_ploted_y[0])
    plt.annotate((to_be_ploted_x[0],float("{0:.3f}".format(to_be_ploted_y[0]))),(to_be_ploted_x[0],to_be_ploted_y[0]))
    if to_be_ploted_y[0]>0.05:
        for i in range(len(to_be_ploted_x)):
            if to_be_ploted_y[i]<=0.05 and alarm_005 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_005 = True
                bar_graph.append("FDR 5%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
                if to_be_ploted_y[i]<=0.01:
                    alarm_001 = True
                    bar_graph[1] = "FDR 1%"
                    #gold.append(correct_no_homo_XL[i])
                    #silver.append(homo_XL[i])
                    #bronze.append(false_XL[i])
            if to_be_ploted_y[i]<=0.01 and alarm_001 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_001 = True
                bar_graph.append("FDR 1%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
    elif to_be_ploted_y[0]<=0.05 and to_be_ploted_y[0]>0.01:
        for i in range(len(to_be_ploted_x)):
            if to_be_ploted_y[i]<=0.01 and alarm_001 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_001 = True
                bar_graph.append("FDR 1%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
        
    
    plt.savefig(name_of_the_image+"plink_FDR.png")
    plt.clf()

    import numpy as np
    
    b_bronze = list (np.add(gold, silver))

    plt.bar(bar_graph,gold,0.3,label="correct XL",color="green")
    plt.bar(bar_graph,silver,0.3,bottom=gold, label="homeotypic",color="cyan")
    plt.bar(bar_graph,bronze,0.3,bottom=b_bronze, label="false",color="red")
    
    plt.xlabel("FDR")
    plt.ylabel("Number of crosslinks")
    plt.title("Type of crosslinks with the FDRCUTOFF="+ fdr_cutoff_value)
    plt.legend()
    plt.savefig(name_of_the_image+"plink_FDR2.png")

    plt.clf()
    plt.xlabel("-log(Score)")
    plt.ylabel("Number of crosslinks")
    plt.stackplot(list_of_scores_plink,correct_no_homo_XL,homo_XL,false_XL,labels=["correct","correct homeotypic","false"], colors=["green","cyan","red"])
    plt.legend()
    plt.savefig(name_of_the_image +"plink_FDR3.png")

fdr_diagamm(unique_crosslinks)










