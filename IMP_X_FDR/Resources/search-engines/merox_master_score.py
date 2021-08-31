#MeroX
import xlsxwriter
import csv
import argparse
# import the necessary library in order to work with xlsx sheets
import xlrd
import os
from zipfile import ZipFile
from openpyxl import Workbook
import csv
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
import csv
import shutil


xlrd.xlsx.ensure_elementtree_imported(False, None)
xlrd.xlsx.Element_has_iter = True

parser = argparse.ArgumentParser()
parser.add_argument("input_file_name")
parser.add_argument("support_file_name")
parser.add_argument("output_file_name")
parser.add_argument("-sintra", "--score_intralink",help="introduce the intraprotein XL cut-off score for the selected FDR. It can be found in -Show decoy analysis-",type=float, default=0)
parser.add_argument("-sinter", "--score_interlink",help="introduce the interprotein XL cut-off score for the selected FDR. It can be found in -Show decoy analysis- ",type=float, default=0)

args = parser.parse_args()



# open the sheetspreadfile and the respective sheet
groups_support_file = args.support_file_name

workbook_groups = xlrd.open_workbook(groups_support_file, on_demand = True)
worksheet_groups = workbook_groups.sheet_by_index(0)

number_groups = 0
peptides_per_group = []
# identify the number of groups
for i in range(worksheet_groups.nrows):
    if worksheet_groups.cell(i, 0).value.find("Group")!=-1:
        number_groups = number_groups +1
        peptides_per_group.append(i)

# add the position where the first empty cell is found
peptides_per_group.append(worksheet_groups.nrows)

#subtract the positions of group headers in order to calculate the number of peptides per group
for i in range(number_groups):
    peptides_per_group[i] = peptides_per_group[i+1]-peptides_per_group[i]-1

peptides_per_group = peptides_per_group[:-1]

print(peptides_per_group)
print(len(peptides_per_group))



#list_crosslinks_connect_acid = ["D","E"]
#list_crosslinks_connect_lys = ["K","Y","T","S"]



score_cutoff_interlink_procent = args.score_interlink
score_cutoff_intralink_procent = args.score_intralink








# in list_of_groups there will be all peptides groups with all their members stored 
list_of_groups = []
# will be added to the list_of_groups and contains temporary the members of one group at the time
list_of_peptides_per_group = []


# all peptide sequences present in the excel document are read one by oneand the added to the list of groups
# list of peptides per group will be permanently emptied to make space for the other groups 
# the number of list of peptides per group is static, not dynamic and should be kept constant
pivot_excel = 1

for i in range(1,number_groups+1):
	if i!=1:
		pivot_excel = pivot_excel+peptides_per_group[i-2]+1
	list_of_peptides_per_group = []
	for j in range(0,peptides_per_group[i-1]):
		list_of_peptides_per_group.append(worksheet_groups.cell(pivot_excel+j, 0).value)
	a = list_of_peptides_per_group
	list_of_groups.append(a)


# take the file with the zhrm extension 
name_file_merox2= args.input_file_name


target = os.path.splitext(args.input_file_name)[0] +"_.zhrm"

shutil.copyfile(name_file_merox2, target)

#convert the file with .zhrm extension into a .zip file in order to make the data accesible

my_file = target
base = os.path.splitext(my_file)[0]
name_of_the_image = base
base = base + '.zip'

os.rename(my_file, base)



  
# specifying the zip file name 
file_name = base
  
# opening the zip file in READ mode 
with ZipFile(file_name, 'r') as zip: 
    # printing all the contents of the zip file 
    zip.printdir() 
  
    # extracting all the files 
    print('Extracting all the files now...')
    for f in zip.filelist:

        if not 'spectra' in f.filename:
            zip.extract(f) 
    print('Done!') 
#we actually don't need the spectra extracted 




#read every line ('counter' is there to keep the counting of each line) from the Result.csv file 
#Result.csv contains all the CSMs
counter=0
wb = Workbook()
ws = wb.active
with open('Result.csv', 'r') as f:
    for row in csv.reader(f, delimiter='§'):
        ws.append(row)
        counter=counter+1
wb.save('Result.xlsx')

wb = xlrd.open_workbook('Result.xlsx', on_demand = True)
ws = wb.sheet_by_name('Sheet')

# needs to be improved
# the problem here is that users can redefine the names of the aminoacids

# character_to_delete contains a list of caracters that are present in the aminoacid sequence but not necessary for analysing the sequence
character_to_delete = ["{","[","]","}","X","J"]

# in csms the following variables are stored
# AA-sequence 1 and 2 
# positions where XL is identified in the peptide ex:"K3"
# the name of both proteins where the peptides come from
# the positions in the proteins where the peptide suquence starts from
#the difference between the csms and the csms_clone is that the score is also stored and in the clone is not
csms = []
csms_clone =[]


#data collection of the characters mentioned above
for i in range(counter):
    pivot = ws.cell(i,0).value.split(";")
    string1 = pivot[6]
    string2 = pivot[10]
    try:
        position1 = pivot[20]
        position2 = pivot[21]

        site_1_protein_name = pivot[7]
        site_2_protein_name = pivot[11]


        # if the the peptide is located right at the beginning of the protein the index needs to be changed in order to mantain the right position
        site_1_starting_position = pivot[8]
        if site_1_starting_position == '0':
            site_1_starting_position = '1'
        site_2_starting_position = pivot[12]
        if site_2_starting_position == '0':
            site_2_starting_position = '1'
    except:
        print ("Fehler\nFehler")
        pivot = ws.cell(i,1).value.split(";")
        position1 = pivot[20-len(ws.cell(i,0).value.split(";"))+1]
        position2 = pivot[21-len(ws.cell(i,0).value.split(";"))+1]

        site_1_protein_name = pivot[7]
        site_2_protein_name = pivot[11]

        site_1_starting_position = pivot[8-len(ws.cell(i,0).value.split(";"))+1]
        site_2_starting_position = pivot[12-len(ws.cell(i,0).value.split(";"))+1]

        
    for character in character_to_delete :
        string1 = string1.replace(character, "")
        string2 = string2.replace(character, "")

    position_take_1 = list(position1)
    position_take_2 = list(position2)

    addition_1 = 0
    addition_2 = 0

    if len(position_take_1)>2:
        j=2
        
        while(j<=len(position_take_1)):
            addition_1 = addition_1*10 + int(position_take_1[j-1])
            j=j+1
        
        total_position_1 = str(addition_1+int(site_1_starting_position)-1)
    else:
        total_position_1 = str(int(position_take_1[1])+int(site_1_starting_position)-1)
    
    if len(position_take_2)>2:
        j=2

        while(j<=len(position_take_2)):
            addition_2 = addition_2*10 + int(position_take_2[j-1])
            j=j+1
        
        total_position_2 = str(addition_2+int(site_2_starting_position)-1)
    else:
        total_position_2 = str(int(position_take_2[1])+int(site_2_starting_position)-1)

    #in case there are two binding aminoacids
    try:
        score_holder = float(pivot[22][0]+pivot[22][1]+pivot[22][2]+pivot[22][3]+pivot[22][4]+pivot[22][5]+pivot[22][6])

        if pivot[22].find("#")!=-1:
            place = pivot[22].find("#")
            comparable = float(pivot[22][0+place+1]+pivot[22][1+place+1]+pivot[22][2+place+1]+pivot[22][3+place+1]+pivot[22][4+place+1]+pivot[22][5+place+1]+pivot[22][6+place+1])
            if comparable>=float(pivot[22][0]+pivot[22][1]+pivot[22][2]+pivot[22][3]+pivot[22][4]+pivot[22][5]+pivot[22][6]):
                score_holder= comparable
            else:
                score_holder= float(pivot[22][0]+pivot[22][1]+pivot[22][2]+pivot[22][3]+pivot[22][4]+pivot[22][5]+pivot[22][6])
            #in case there are three binding aminoacids
            if pivot[22].find("#", place+1)!=-1:
                place2 = pivot[22].find("#", place+1)
                comparable = float(pivot[22][0+place2+1]+pivot[22][1+place2+1]+pivot[22][2+place2+1]+pivot[22][3+place2+1]+pivot[22][4+place2+1]+pivot[22][5+place2+1]+pivot[22][6+place2+1])
                if comparable>=score_holder:
                    score_holder= comparable
    except:
        score_holder = float(pivot[0])






    if site_1_protein_name==site_2_protein_name:
        if float(score_holder)>=score_cutoff_intralink_procent:
            csms.append([string1,string2,position1,position2,site_1_protein_name,site_2_protein_name,site_1_starting_position,site_2_starting_position,total_position_1,total_position_2,pivot[0]])
            csms_clone.append([string1,string2,position1,position2,site_1_protein_name,site_2_protein_name,site_1_starting_position,site_2_starting_position,total_position_1,total_position_2])
    elif float(score_holder)>=score_cutoff_interlink_procent:
            csms.append([string1,string2,position1,position2,site_1_protein_name,site_2_protein_name,site_1_starting_position,site_2_starting_position,total_position_1,total_position_2,pivot[0]])
            csms_clone.append([string1,string2,position1,position2,site_1_protein_name,site_2_protein_name,site_1_starting_position,site_2_starting_position,total_position_1,total_position_2])
total_number_of_CSMs = len(csms)






# this function removes duplicates and returns a tuple
# not  effective for csms bc same crosslinks could have different scores
# the reason why clones need to be created the whole time 
# manageble bc the lists are normally small
def removeDuplicates(lst): 
      
    return [t for t in (set(tuple(i) for i in lst))]



modification_dictionary ={}
fdr_cutoff_value = ''

def repare_modifications():
    
    
    #aminoacids_list = ('A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'V', 'W','Y')
    aminoacids_list_temporary = []


    # in the file properties.mxf the dictionaries for the modifications and for the aminoacids are stored 
    with open("properties.mxf") as file_mxf:
        lines_mxf = file_mxf.readlines()
    for i in range(len(lines_mxf)-1,0,-1):
        if lines_mxf[i].find('AMINOACIDS')!=-1:
            j = i+1
            while lines_mxf[j].find('END')==-1:
                a = lines_mxf[j].find(';')
                aminoacids_list_temporary.append(lines_mxf[j][a+1])
                j = j+1
    aminoacids_list = tuple(aminoacids_list_temporary)
    # aminoacid_list looks like that = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'V', 'W', 'X', 'Y', '[', ']', '#', '(', 'k', 'm', '/', '{', '}')
    # so it does include all the possible modifications as well


    for i in range(len(lines_mxf)-1,0,-1):
        if lines_mxf[i].find('VARMODIFICATION')!=-1:
            j=i+1
            while lines_mxf[j].find('END')==-1:
                if lines_mxf[j][2] in aminoacids_list:
                    modification_dictionary[lines_mxf[j][2]]=lines_mxf[j][4]
                j=j+1
    # in this dictionary there are the translations of the existing modifcations "modification_dictionary"
    for i in range(len(lines_mxf)-1,0,-1):
        if lines_mxf[i].find('STATMODIFICATION')!=-1:
            j=i+1
            while lines_mxf[j].find('END')==-1:
                if lines_mxf[j][2] in aminoacids_list:
                    modification_dictionary[lines_mxf[j][2]]=lines_mxf[j][4]
                j=j+1
    
    find_fdr_cutoff = False
    i = 0
    while find_fdr_cutoff == False and i<len(lines_mxf):
        if lines_mxf[i].find('FDRCUTOFF=')!=-1:
            fdr_cutoff_value = str(lines_mxf[i][10:14])
            find_fdr_cutoff = True
        else:
            i=i+1
    return fdr_cutoff_value


    print(aminoacids_list)


# function to return key for any value 
def get_key(val,my_dict): 
    for key, value in my_dict.items(): 
         if val == value: 
             return key 
  
    return "key doesn't exist"
     



# the call of reparing the modifications
# m -> M, oxidised methionine to methionine for example
# VERY IMPORTANT ---- this functions just accumulates the data in order to do the modifications but doesn't "repair" the sequences

fdr_cutoff_value = repare_modifications()



for i in range(len(csms)):
    site1 = list(csms[i][0])
    site2 = list(csms[i][1]) 
    for j in range(len(csms[i][0])):
        if site1[j] in modification_dictionary.values():
            site1[j] = get_key(site1[j],modification_dictionary)
            
    
    csms[i][0]= ''.join(site1)
    csms_clone[i][0] = ''.join(site1)
    for j in range(len(csms[i][1])):
        if site2[j] in modification_dictionary.values():
            site2[j]= get_key(site2[j],modification_dictionary)
    csms[i][1]= ''.join(site2)
    csms_clone[i][1]=''.join(site2)

    for character in character_to_delete :
        csms[i][0] = csms[i][0].replace(character, "")
        csms[i][1] = csms[i][1].replace(character, "")
        csms_clone[i][0] = csms_clone[i][0].replace(character, "")
        csms_clone[i][1] = csms_clone[i][1].replace(character, "")
    
#the code from above "repairs" the modifications


def order_alphabetically(example_list):

    example_list2= []

    for i in range(len(example_list)):
        example_list2 = example_list[i][0:2]

        
        if (example_list[i][0:2] != sorted(example_list2[0:2])):
            
            exchange = example_list[i][0]
            example_list[i][0] = example_list[i][1]
            example_list[i][1] = exchange

            exchange = example_list[i][2]
            example_list[i][2] = example_list[i][3]
            example_list[i][3] = exchange

            exchange = example_list[i][4]
            example_list[i][4] = example_list[i][5]
            example_list[i][5] = exchange

            exchange = example_list[i][6]
            example_list[i][6] = example_list[i][7]
            example_list[i][7] = exchange

            exchange = example_list[i][8]
            example_list[i][8] = example_list[i][9]
            example_list[i][9] = exchange

    return  example_list;



csms = order_alphabetically(csms)
csms_clone = order_alphabetically(csms_clone)




csms2 = []
csms_clone2 = []
csms3 = []
csms_clone3 = []
csms4 = []


csms2 = removeDuplicates(csms)
csms_clone2 = removeDuplicates(csms_clone) 

# we remove now the duplicates from csms and csms_clone
# now they have a different number of values




name_and_position_of_the_pair = []
list_of_scores = []
count=0

#lösche danach bitte
temporarylist = []

# the dead-end crosslinks are now eliminated from the csms_clone2
for i in range(len(csms_clone2)):
    if(csms_clone2[i][0] !='0' and csms_clone2[i][1]!='0' and csms_clone2[i][0] !='1' and csms_clone2[i][1]!='1'):
        count = count+1
        csms_clone3.append(csms_clone2[i])
        
        
        name_and_position_of_the_pair.append([csms_clone2[i][4],csms_clone2[i][5],csms_clone2[i][8],csms_clone2[i][9]])
        #lösche danach bitte
        temporarylist.append([str(csms_clone2[i][4]+csms_clone2[i][8]),str(csms_clone2[i][5]+csms_clone2[i][9])])
        #print(str(csms_clone2[i][4]+csms_clone2[i][8]+csms_clone2[i][5]+csms_clone2[i][9]))
        #print(count)

good_count = count


#lösche danach bitte
for i in range(good_count):
    temporarylist[i].sort()
    print(temporarylist[i][0]+temporarylist[i][1])


dictionary_of_best_score = {}


count=0

# the dead-end crosslinks are now eliminated from the csms2
for i in range(len(csms2)):
    if(csms2[i][0] !='0' and csms2[i][1]!='0' and csms2[i][0] !='1' and csms2[i][1]!='1' ):
        count = count+1
        csms3.append(csms2[i])
        
        pivot_helper = tuple(csms2[i][:-1])
        if(pivot_helper in dictionary_of_best_score.keys()):
            if int(csms2[i][10])> dictionary_of_best_score[pivot_helper]:
                dictionary_of_best_score[pivot_helper] = int(csms2[i][10])
        else:
            dictionary_of_best_score[pivot_helper]= int(csms2[i][10])
        #additionally we save the score or the best score(in case of duplicates) for each crosslink     
        
        
        list_of_scores.append(int(csms2[i][10]))
        # it is not a problem if we have scores of the crosslinks that we later eliminate
        # this could also have been done from a theoretically generated score list(from 0 to the highest score)

j=0
k=0


#temp2 = []
#temp1 = []



print(list_of_scores)
#eliminate the duplicates
list_of_scores = list(set(list_of_scores))
#sort the list of scores
list_of_scores.sort()

def fdr_diagamm(list_crosslinks):

    temp2 = []
    temp1 = []

    # Create a workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(os.path.splitext(args.output_file_name)[0] + '_venn_input.xlsx')
    worksheet = workbook.add_worksheet()

    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0

    worksheet.write(0,0, "Results")
    worksheet.write(2,0, "Total number of CSMs:")
    worksheet.write(3,0, "Total number of unique XLs:")
    worksheet.write(4,0, "All True crosslinks: ")
    worksheet.write(5,0, "Homeotypic crosslinks:")
    worksheet.write(6,0, "Non-homeotypic crosslinks:")
    worksheet.write(7,0, "False crosslinks")

    worksheet.write(9,0, "All crosslinks")
    worksheet.write(9,1, "True crosslinks")
    worksheet.write(9,2, "False crosslinks")

    worksheet.write(2,3,"number of XLs post-score cut-off 5%:")
    worksheet.write(3,3,"numebr of XLs post-score cut-off 1%:")

    worksheet.set_column(0,0,40)
    worksheet.set_column(0,1,40)
    worksheet.set_column(0,2,40)
    worksheet.set_column(0,3,40)

    
    # open the file in the write mode
    f = open(args.output_file_name, 'w', newline='')

    #define the header
    header_csv = ["Sequence A", "Sequence B","Accession A","Accession B","Position in protein A","Position in protein B","Score crosslink","Within same group"]

    # create the csv writer
    writer = csv.writer(f)
    writer.writerow(header_csv)

    list_true_XL_csv = []
    list_false_XL_csv = []

    


    to_be_ploted_x = []
    to_be_ploted_y = []
    correct_no_homo_XL = []
    homo_XL = []
    false_XL = []



    for i in range(len(list_of_scores)):
        all_crosslinks = 0
        correct_crosslinks = 0
        homeotypic = 0
        for j in range(len(list_crosslinks)):
            if list_crosslinks[list_dictionary_of_the_best_score[j]]>=list_of_scores[i]:
                all_crosslinks = all_crosslinks +1
                temp1.append(list([list_dictionary_of_the_best_score[j][0],list_dictionary_of_the_best_score[j][1],
                                   list_dictionary_of_the_best_score[j][4],list_dictionary_of_the_best_score[j][5],
                                   list_dictionary_of_the_best_score[j][8],list_dictionary_of_the_best_score[j][9], str("score_" + str(list_crosslinks[list_dictionary_of_the_best_score[j]]))]))
                found = False
                k = 0
                while(k<number_groups and found == False):
                    l = 0
                    while(l<peptides_per_group[k] and found == False):
                        if list_dictionary_of_the_best_score[j][0] == list_of_groups[k][l] or list_dictionary_of_the_best_score[j][0] in list_of_groups[k][l] :
                            m = 0
                            while(m<peptides_per_group[k] and found== False):
                                if (list_dictionary_of_the_best_score[j][1]== list_of_groups[k][m] or list_dictionary_of_the_best_score[j][1] in list_of_groups[k][m]):
                                    found = True
                                    correct_crosslinks = correct_crosslinks +1
                                    temp2.append(list([list_dictionary_of_the_best_score[j][0],list_dictionary_of_the_best_score[j][1],
                                                       list_dictionary_of_the_best_score[j][4],list_dictionary_of_the_best_score[j][5],
                                                       list_dictionary_of_the_best_score[j][8],list_dictionary_of_the_best_score[j][9], str("score_" + str(list_crosslinks[list_dictionary_of_the_best_score[j]]))]))
                                    if i==0:
                                        list_true_XL_csv.append([list_dictionary_of_the_best_score[j][0],list_dictionary_of_the_best_score[j][1],
                                                       list_dictionary_of_the_best_score[j][4],list_dictionary_of_the_best_score[j][5],
                                                       list_dictionary_of_the_best_score[j][8],list_dictionary_of_the_best_score[j][9], list_crosslinks[list_dictionary_of_the_best_score[j]],"TRUE"])

                                    if list_dictionary_of_the_best_score[j][0] == list_dictionary_of_the_best_score[j][1]:
                                        homeotypic = homeotypic+1
                                else:
                                    m = m+1
                            
                        l = l+1
                    k = k+1
                if i==0 and found == False:
                    list_false_XL_csv.append([list_dictionary_of_the_best_score[j][0],list_dictionary_of_the_best_score[j][1],
                                                       list_dictionary_of_the_best_score[j][4],list_dictionary_of_the_best_score[j][5],
                                                       list_dictionary_of_the_best_score[j][8],list_dictionary_of_the_best_score[j][9], list_crosslinks[list_dictionary_of_the_best_score[j]],"FALSE"])

        if i == 0:

            def working_on_name_of_the_protein(list_of_lists):
                for i in range(len(list_of_lists)):
                    j = list_of_lists[i][2].find("sg|")+3
                    k = list_of_lists[i][3].find("sg|")+3
        
                    little_pivot1 = list(list_of_lists[i][2])
                    little_pivot2 = list(list_of_lists[i][3])
                    SG1_name = ""
                    SG2_name = "" 
                    while(little_pivot1[j]!='|'):
                        SG1_name = SG1_name + little_pivot1[j]
                        j=j+1
                    while(little_pivot2[k]!='|'):
                        SG2_name = SG2_name + little_pivot2[k]
                        k=k+1

                    list_of_lists[i][2]=SG1_name
                    list_of_lists[i][3]=SG2_name

                return list_of_lists
            try:
                temp1 = working_on_name_of_the_protein(temp1)
                temp2 = working_on_name_of_the_protein(temp2)
            except:
                print("name of the protein not appropriate!")
            for m in range(all_crosslinks):
                temp1[m].sort()
                temp1[m]=str(temp1[m])
                worksheet.write(11+m,0,str(temp1[m]))

            for m in range(len(temp2)):
                temp2[m].sort()
                temp2[m]=str(temp2[m])
                worksheet.write(11+m,1,str(temp2[m]))

            temp3=list(set(temp1)-set(temp2)) 
            for m in range(len(temp3)):
                worksheet.write(11+m,2,str(temp3[m]))
                



            list_of_the_correct_crosslinks = set(temp2)
            list_of_the_false_crosslinks = set(set(temp1)-set(temp2))
            list_of_all_crosslinks = set(temp1)


            



        to_be_ploted_x.append(list_of_scores[i])
        to_be_ploted_y.append((all_crosslinks-correct_crosslinks)/all_crosslinks)
        correct_no_homo_XL.append(correct_crosslinks-homeotypic)
        homo_XL.append(homeotypic)
        false_XL.append(all_crosslinks-correct_crosslinks)
        
    
    plt.plot(to_be_ploted_x,to_be_ploted_y)

    plt.title('Real FDR dependent on the score')
    plt.xlabel('Score')
    plt.ylabel('FDR')
    alarm_005 = False
    alarm_001 = False

    try:
        worksheet.write(2,1,total_number_of_CSMs)
        worksheet.write(3,1,correct_no_homo_XL[0]+homo_XL[0]+false_XL[0])
        worksheet.write(4,1,correct_no_homo_XL[0]+homo_XL[0])
        worksheet.write(5,1,homo_XL[0])
        worksheet.write(6,1,correct_no_homo_XL[0])
        worksheet.write(7,1,false_XL[0])
    except:
        print("ATTENZIONE")
        print("ATTENZIONE")

    
    bar_graph = ["real FDR"+" " +str(float("{0:.1f}".format((false_XL[0]/(correct_no_homo_XL[0]+homo_XL[0]+false_XL[0]))*100)))+"%"]
    
    gold = [correct_no_homo_XL[0]]
    silver = [homo_XL[0]]
    bronze = [false_XL[0]]
    score_bar = [to_be_ploted_x[0]]


    plt.scatter(to_be_ploted_x[0],to_be_ploted_y[0])
    plt.annotate((to_be_ploted_x[0],float("{0:.3f}".format(to_be_ploted_y[0]))),(to_be_ploted_x[0],to_be_ploted_y[0]))
    if to_be_ploted_y[0]>0.05:
        for i in range(len(to_be_ploted_x)):
            if to_be_ploted_y[i]<=0.05 and alarm_005 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_005 = True
                worksheet.write(2,4,correct_no_homo_XL[i]+homo_XL[i]+false_XL[i])
                bar_graph.append("FDR 5%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
                score_bar.append(to_be_ploted_x[i])
                if to_be_ploted_y[i]<=0.01:
                    alarm_001 = True
                    worksheet.write(3,4,correct_no_homo_XL[i]+homo_XL[i]+false_XL[i])
                    bar_graph[1] = "FDR 1%"
                    #gold.append(correct_no_homo_XL[i])
                    #silver.append(homo_XL[i])
                    #bronze.append(false_XL[i])
            if to_be_ploted_y[i]<=0.01 and alarm_001 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_001 = True
                worksheet.write(3,4,correct_no_homo_XL[i]+homo_XL[i]+false_XL[i])
                bar_graph.append("FDR 1%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
                score_bar.append(to_be_ploted_x[i])
    elif to_be_ploted_y[0]<=0.05 and to_be_ploted_y[0]>0.01:
        for i in range(len(to_be_ploted_x)):
            if to_be_ploted_y[i]<=0.01 and alarm_001 == False:
                plt.scatter(to_be_ploted_x[i],to_be_ploted_y[i])
                plt.annotate((to_be_ploted_x[i],float("{0:.3f}".format(to_be_ploted_y[i]))),(to_be_ploted_x[i],to_be_ploted_y[i]))
                alarm_001 = True
                worksheet.write(3,4,correct_no_homo_XL[i]+homo_XL[i]+false_XL[i])
                bar_graph.append("FDR 1%")
                gold.append(correct_no_homo_XL[i])
                silver.append(homo_XL[i])
                bronze.append(false_XL[i])
                score_bar.append(to_be_ploted_x[i])
        
    workbook.close()

    plt.savefig(os.path.splitext(args.output_file_name)[0] + "_ScorevsFDR.svg")
    plt.clf()

    
    b_bronze = list (np.add(gold, silver))

    plt.bar(bar_graph,gold,label="true",color="green")
    plt.bar(bar_graph,silver,bottom=gold, label="true homeotypic",color="lawngreen")
    plt.bar(bar_graph,bronze,bottom=b_bronze, label="false",color="red")
    
    plt.ylabel("Number of crosslinks")
    plt.title("FDR-CUT-OFF-SCORE ="+ fdr_cutoff_value)
    plt.legend()
    plt.tick_params(
    axis='x',
    which='both',
    bottom=False,
    top=False,
    labelbottom=False)

    cell_table = []
    cell_table.append(gold)
    cell_table.append(silver)
    cell_table.append(bronze)
    cell_table.append(score_bar)
    plt.table(cellText=cell_table, cellLoc='center', rowLabels=["true","true homeotypic","false","at score"],colLabels=bar_graph ,rowLoc='left' ,colLoc='center', loc='bottom', edges='closed')
    

    plt.savefig(os.path.splitext(args.output_file_name)[0] + "_numberXLs.svg",bbox_inches = "tight")

    plt.clf()
    plt.xlabel("Score")
    plt.ylabel("Number of crosslinks")
    plt.stackplot(list_of_scores,correct_no_homo_XL,homo_XL,false_XL,labels=["true","true homeotypic","false"], colors=["green","lawngreen","red"])
    plt.legend()
    plt.savefig(os.path.splitext(args.output_file_name)[0] + "_ScorevsNumber.svg")

    writer.writerows(list_true_XL_csv)
    writer.writerows(list_false_XL_csv)
    f.close()
#
#
list_dictionary_of_the_best_score = list(dictionary_of_best_score)
print(len(list_dictionary_of_the_best_score))
print(dictionary_of_best_score)
#fdr_diagamm(dictionary_of_best_score)
                    
                

    


for i in range(good_count):
    found = False
    j=0
    while (j<number_groups and found == False):
        k=0
        while (k<peptides_per_group[j] and found == False ):
            if (csms_clone3[i][0]==list_of_groups[j][k] or csms_clone3[i][0] in list_of_groups[j][k]):
                l=0
                while (l<peptides_per_group[j] and found == False):
                    if (csms_clone3[i][1]==list_of_groups[j][l] or csms_clone3[i][1] in list_of_groups[j][l]):
                        found = True
                        csms4.append((list_of_groups[j][k],list_of_groups[j][l]))
                        
                    else:
                        l=l+1
            
            k=k+1
        else:
            j=j+1



#for i in range(len(csms4)):
   # print(csms4[i])

print(str(len(csms4)) +" out of "+ str(len(csms_clone3)) +" are correct")
try:
    # just for testing, will be erased later
    for i in range(len(name_and_position_of_the_pair)):
        j = name_and_position_of_the_pair[i][0].find("GN=")+3
        k = name_and_position_of_the_pair[i][1].find("GN=")+3
        
        little_pivot1 = list(name_and_position_of_the_pair[i][0])
        little_pivot2 = list(name_and_position_of_the_pair[i][1])
        GN1_name = ""
        GN2_name = "" 
        while(little_pivot1[j]!=' '):
            GN1_name = GN1_name + little_pivot1[j]
            j=j+1
        while(little_pivot2[k]!=' '):
            GN2_name = GN2_name + little_pivot2[k]
            k=k+1
except:
    pass
    #
    # print(GN1_name, name_and_position_of_the_pair[i][2],GN2_name,name_and_position_of_the_pair[i][3])
    mini_list = []
    mini_list.append(GN1_name + str(name_and_position_of_the_pair[i][2]))
    mini_list.append(GN2_name + str(name_and_position_of_the_pair[i][3]))
    mini_list.sort()
    print(mini_list[0],mini_list[1])


fdr_diagamm(dictionary_of_best_score)

print(len(dictionary_of_best_score),good_count)

try:
    os.remove("coreResult.csv")
except:
    pass

try:
    os.remove("decoy.csv")
except:
    pass

try:
    os.remove("properties.mxf")
except:
    pass

try:
    os.remove("proteinLength.csv")
except:
    pass

try:
    os.remove("report.txt")
except:
    pass

try:
    os.remove("Result.csv")
except:
    pass

try:
    os.remove("Result.xlsx")
except:
    pass

try:
    os.remove("db.fasta")
except:
    pass


target = os.path.splitext(args.input_file_name)[0] + "_.zip"

os.remove(target)
